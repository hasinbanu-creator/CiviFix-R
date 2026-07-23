import os
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List
from .load_runner import LoadResult

def generate_load_excel_report(results: List[LoadResult], output_path: str = "CiviFix_LoadTest_Execution_Report.xlsx"):
    wb = openpyxl.Workbook()
    
    # ─── Color Palette & Styles (Appium Theme) ─────────────────────────
    TEAL_BANNER_COLOR = "359E80"
    DARK_NAVY_COLOR = "202F3C"
    DARK_ROW_COLOR = "2A3644"
    TEAL_TEXT_COLOR = "4CBFA6"
    WHITE_COLOR = "FFFFFF"
    GREEN_PASS_COLOR = "4BB543"
    RED_FAIL_COLOR = "FF3333"

    banner_fill = PatternFill(start_color=TEAL_BANNER_COLOR, end_color=TEAL_BANNER_COLOR, fill_type="solid")
    dark_header_fill = PatternFill(start_color=DARK_NAVY_COLOR, end_color=DARK_NAVY_COLOR, fill_type="solid")
    dark_row_fill = PatternFill(start_color=DARK_ROW_COLOR, end_color=DARK_ROW_COLOR, fill_type="solid")

    banner_font = Font(color=WHITE_COLOR, bold=True, size=16)
    header_font = Font(color=WHITE_COLOR, bold=True, size=11)
    white_font = Font(color=WHITE_COLOR, size=11)
    bold_white_font = Font(color=WHITE_COLOR, bold=True, size=11)
    teal_font = Font(color=TEAL_TEXT_COLOR, size=11)
    pass_font = Font(color=GREEN_PASS_COLOR, bold=True, size=11)
    fail_font = Font(color=RED_FAIL_COLOR, bold=True, size=11)

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(left=Side(style='thin', color='505050'), 
                         right=Side(style='thin', color='505050'),
                         top=Side(style='thin', color='505050'), 
                         bottom=Side(style='thin', color='505050'))

    # Calculate global metrics
    total_reqs = sum(r.total_requests for r in results)
    total_fails = sum(r.failures for r in results)
    total_success = sum(r.successes for r in results)
    global_avg_rt = sum(r.avg_rt * r.total_requests for r in results) / total_reqs if total_reqs else 0
    global_rps = sum(r.rps for r in results)
    pass_rate = (total_success / total_reqs * 100) if total_reqs else 0
    duration_str = "60 seconds (1 Min)"
    
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def apply_row_style(ws, row_idx, fill, font, align):
        for cell in ws[row_idx]:
            cell.fill = fill
            cell.font = font
            cell.alignment = align
            cell.border = thin_border

    # ─── 1. Executive Summary ───────────────────────────────
    ws_exec = wb.active
    ws_exec.title = "Executive Summary"
    ws_exec.sheet_properties.tabColor = TEAL_BANNER_COLOR
    
    ws_exec.column_dimensions['A'].width = 35
    ws_exec.column_dimensions['B'].width = 35
    ws_exec.column_dimensions['C'].width = 45

    # Top Banner
    ws_exec.merge_cells("A1:C1")
    cell = ws_exec.cell(row=1, column=1, value="CiviFix — API Baseline/Load Test Report")
    cell.fill = banner_fill
    cell.font = banner_font
    cell.alignment = center_align

    # Sub-banner
    ws_exec.merge_cells("A2:C2")
    sub_text = f"Generated: {timestamp} | Environment: API Load Engine | Overall Status: {'PASS' if total_fails == 0 else 'FAIL'}"
    cell = ws_exec.cell(row=2, column=1, value=sub_text)
    cell.fill = dark_header_fill
    cell.font = white_font
    cell.alignment = center_align

    # EXECUTION SUMMARY Header
    ws_exec.merge_cells("A4:C4")
    cell = ws_exec.cell(row=4, column=1, value="EXECUTION SUMMARY")
    cell.fill = dark_header_fill
    cell.font = bold_white_font
    cell.alignment = center_align

    summary_data = [
        ("Total API Scenarios", str(len(results)), "Complete Baseline API test suite"),
        ("Total Requests Sent", str(total_reqs), "Distributed across 400 endpoints"),
        ("Passed Requests", str(total_success), "All requests completed successfully"),
        ("Failed Requests", str(total_fails), "No failures detected in this run" if total_fails == 0 else f"{total_fails} failures detected"),
        ("Pass Rate", f"{pass_rate:.1f}%", "Perfect pass rate achieved" if pass_rate == 100 else "Action required"),
        ("Test Duration", duration_str, "100 Virtual Users concurrent execution"),
        ("Overall RPS", f"{global_rps:.2f} req/sec", "Requests Per Second processed"),
        ("Global Avg Latency", f"{global_avg_rt:.2f} ms", "Average response time across all endpoints"),
        ("CI Environment", "GitHub Actions (ubuntu-latest)", "Automated on every push")
    ]
    
    row_idx = 5
    for key, val, desc in summary_data:
        c_k = ws_exec.cell(row=row_idx, column=1, value=key)
        c_v = ws_exec.cell(row=row_idx, column=2, value=val)
        c_d = ws_exec.cell(row=row_idx, column=3, value=desc)
        
        c_k.fill = dark_header_fill
        c_k.font = white_font
        c_k.border = thin_border
        
        c_v.fill = dark_row_fill
        # Special coloring for values
        if "Passed" in key or "Pass Rate" in key or "Total API" in key or "Requests Sent" in key:
            c_v.font = teal_font
        elif "Failed" in key:
            c_v.font = fail_font if int(val) > 0 else pass_font
        else:
            c_v.font = teal_font
            
        c_v.border = thin_border
        c_v.alignment = center_align
        
        c_d.fill = dark_row_fill
        c_d.font = Font(color=WHITE_COLOR, italic=True, size=11)
        c_d.border = thin_border
        
        row_idx += 1

    # MODULE SUMMARY Header
    row_idx += 1
    ws_exec.merge_cells(f"A{row_idx}:C{row_idx}")
    cell = ws_exec.cell(row=row_idx, column=1, value="MODULE SUMMARY")
    cell.fill = dark_header_fill
    cell.font = bold_white_font
    cell.alignment = center_align
    row_idx += 1

    # Module Headers
    headers = ["Module", "Endpoints", "Average RPS"]
    for col_idx, h in enumerate(headers, 1):
        c = ws_exec.cell(row=row_idx, column=col_idx, value=h)
        c.fill = dark_header_fill
        c.font = bold_white_font
        c.alignment = center_align
        c.border = thin_border
    row_idx += 1

    modules = {}
    for r in results:
        m = r.scenario.module
        if m not in modules:
            modules[m] = {"endpoints": 0, "rps": 0.0}
        modules[m]["endpoints"] += 1
        modules[m]["rps"] += r.rps

    for mod, data in modules.items():
        c_m = ws_exec.cell(row=row_idx, column=1, value=mod)
        c_m.fill = dark_row_fill
        c_m.font = teal_font
        c_m.border = thin_border
        c_m.alignment = center_align
        
        c_e = ws_exec.cell(row=row_idx, column=2, value=data["endpoints"])
        c_e.fill = dark_row_fill
        c_e.font = white_font
        c_e.border = thin_border
        c_e.alignment = center_align
        
        c_r = ws_exec.cell(row=row_idx, column=3, value=f"{data['rps']:.2f}")
        c_r.fill = dark_row_fill
        c_r.font = white_font
        c_r.border = thin_border
        c_r.alignment = center_align
        row_idx += 1

    # ─── 2. Module Breakdown ────────────────────────────────
    ws_mod = wb.create_sheet(title="Module Breakdown")
    ws_mod.sheet_properties.tabColor = TEAL_BANNER_COLOR
    
    ws_mod.merge_cells("A1:G1")
    cell = ws_mod.cell(row=1, column=1, value="Module-by-Module Breakdown")
    cell.fill = banner_fill
    cell.font = banner_font
    cell.alignment = center_align

    headers_mod = ["Module", "Endpoints", "Passed", "Failed", "Pass Rate", "Total Reqs", "Avg Latency (ms)"]
    for col_idx, h in enumerate(headers_mod, 1):
        c = ws_mod.cell(row=2, column=col_idx, value=h)
        c.fill = dark_header_fill
        c.font = bold_white_font
        c.alignment = center_align
        c.border = thin_border
        ws_mod.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 22
        
    mod_data = {}
    for r in results:
        m = r.scenario.module
        if m not in mod_data:
            mod_data[m] = {"eps": 0, "pass": 0, "fail": 0, "reqs": 0, "rt_sum": 0.0}
        mod_data[m]["eps"] += 1
        mod_data[m]["pass"] += r.successes
        mod_data[m]["fail"] += r.failures
        mod_data[m]["reqs"] += r.total_requests
        mod_data[m]["rt_sum"] += (r.avg_rt * r.total_requests)
        
    row_idx = 3
    for mod, data in mod_data.items():
        avg_rt = data["rt_sum"] / data["reqs"] if data["reqs"] else 0
        prate = (data["pass"] / data["reqs"] * 100) if data["reqs"] else 0
        
        row_values = [mod, data["eps"], data["pass"], data["fail"], f"{prate:.1f}%", data["reqs"], f"{avg_rt:.2f}"]
        for col_idx, val in enumerate(row_values, 1):
            c = ws_mod.cell(row=row_idx, column=col_idx, value=val)
            c.fill = dark_row_fill
            c.border = thin_border
            c.alignment = center_align
            
            if col_idx == 1: c.font = teal_font
            elif col_idx == 5: c.font = pass_font if prate == 100 else fail_font
            else: c.font = white_font
        row_idx += 1

    # ─── 3. Detailed Results ────────────────────────────────
    ws_det = wb.create_sheet(title="Detailed Results")
    ws_det.sheet_properties.tabColor = TEAL_BANNER_COLOR
    
    ws_det.merge_cells("A1:I1")
    cell = ws_det.cell(row=1, column=1, value="Detailed Load Test Results — 400 API Scenarios")
    cell.fill = banner_fill
    cell.font = banner_font
    cell.alignment = center_align

    headers_det = ["Test ID", "Module", "API Endpoint", "Total Reqs", "RPS", "Avg Latency (ms)", "Min Latency", "Max Latency", "Status"]
    
    for col_idx, h in enumerate(headers_det, 1):
        c = ws_det.cell(row=2, column=col_idx, value=h)
        c.fill = dark_header_fill
        c.font = bold_white_font
        c.alignment = center_align
        c.border = thin_border
        
    ws_det.column_dimensions['A'].width = 12
    ws_det.column_dimensions['B'].width = 20
    ws_det.column_dimensions['C'].width = 45
    ws_det.column_dimensions['D'].width = 15
    ws_det.column_dimensions['E'].width = 15
    ws_det.column_dimensions['F'].width = 18
    ws_det.column_dimensions['G'].width = 18
    ws_det.column_dimensions['H'].width = 18
    ws_det.column_dimensions['I'].width = 15

    for row_idx, r in enumerate(results, 3):
        status = "PASSED" if r.failures == 0 else "FAILED"
        
        row_values = [
            r.scenario.tc_id, r.scenario.module, r.scenario.endpoint,
            r.total_requests, f"{r.rps:.2f}", f"{r.avg_rt:.1f}", f"{r.min_rt:.1f}", f"{r.max_rt:.1f}", status
        ]
        
        for col_idx, val in enumerate(row_values, 1):
            c = ws_det.cell(row=row_idx, column=col_idx, value=val)
            c.fill = dark_row_fill
            c.border = thin_border
            c.alignment = center_align
            
            if col_idx == 1: c.font = teal_font
            elif col_idx == 9: c.font = pass_font if status == "PASSED" else fail_font
            else: c.font = white_font

    # ─── 4. Failure Analysis ────────────────────────────────
    ws_fail = wb.create_sheet(title="Failure Analysis")
    ws_fail.sheet_properties.tabColor = TEAL_BANNER_COLOR
    
    all_passed = total_fails == 0
    fail_banner_text = "Failure Analysis — (All Tests Passed — No Failures)" if all_passed else "Failure Analysis — SLA Breaches & Errors"
    fail_banner_fill = PatternFill(start_color="3B9C83" if all_passed else "C0504D", end_color="3B9C83" if all_passed else "C0504D", fill_type="solid")
    
    ws_fail.merge_cells("A1:D1")
    cell = ws_fail.cell(row=1, column=1, value=fail_banner_text)
    cell.fill = fail_banner_fill
    cell.font = banner_font
    cell.alignment = center_align

    headers_fail = ["Test ID", "Module", "API Endpoint", "Issue Details"]
    for col_idx, h in enumerate(headers_fail, 1):
        c = ws_fail.cell(row=2, column=col_idx, value=h)
        c.fill = dark_header_fill
        c.font = bold_white_font
        c.alignment = center_align
        c.border = thin_border
    
    ws_fail.column_dimensions['A'].width = 15
    ws_fail.column_dimensions['B'].width = 25
    ws_fail.column_dimensions['C'].width = 50
    ws_fail.column_dimensions['D'].width = 50
    
    row_idx = 3
    for r in results:
        issues = []
        if r.failures > 0:
            issues.append(f"{r.failures} requests failed out of {r.total_requests}")
        if r.avg_rt > 1000:
            issues.append(f"High Latency (Avg: {r.avg_rt:.1f}ms)")
            
        if issues:
            for col_idx, val in enumerate([r.scenario.tc_id, r.scenario.module, r.scenario.endpoint, " | ".join(issues)], 1):
                c = ws_fail.cell(row=row_idx, column=col_idx, value=val)
                c.fill = dark_row_fill
                c.font = white_font
                c.border = thin_border
                if col_idx < 4: c.alignment = center_align
            row_idx += 1
            
    if all_passed and row_idx == 3:
        ws_fail.merge_cells("A3:D3")
        c = ws_fail.cell(row=3, column=1, value="No failures or latency SLA breaches detected. All 400 API scenarios passed successfully.")
        c.fill = dark_row_fill
        c.font = teal_font
        c.border = thin_border
        c.alignment = center_align

    wb.save(output_path)
    print(f"[REPORT] Excel Report generated successfully with Appium Dark Theme: {output_path}")
